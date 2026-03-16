"""Tests for Excel output: sheet names, required columns, color legend, data types."""
import sys
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook

import portfolio_constructor
from portfolio_constructor import (
    write_factor_scores_sheet,
    write_screener_dashboard,
    write_model_portfolio_sheet,
    write_data_validation_sheet,
    write_full_excel,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _make_scored_universe(n: int = 20, seed: int = 42) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    sectors = ["Information Technology", "Financials", "Health Care",
               "Consumer Discretionary", "Industrials"]
    df = pd.DataFrame({
        "Ticker": [f"T{i:03d}" for i in range(n)],
        "Company": [f"Company {i}" for i in range(n)],
        "Sector": [sectors[i % len(sectors)] for i in range(n)],
        "valuation_score":  rng.uniform(10, 90, n),
        "quality_score":    rng.uniform(10, 90, n),
        "growth_score":     rng.uniform(10, 90, n),
        "momentum_score":   rng.uniform(10, 90, n),
        "risk_score":       rng.uniform(10, 90, n),
        "revisions_score":  rng.uniform(10, 90, n),
        "size_score":       rng.uniform(10, 90, n),
        "investment_score": rng.uniform(10, 90, n),
        "Composite":        rng.uniform(30, 95, n),
        "Rank":             list(range(1, n + 1)),
        "valuation_contrib":   rng.uniform(0, 10, n),
        "quality_contrib":     rng.uniform(0, 10, n),
        "growth_contrib":      rng.uniform(0, 10, n),
        "momentum_contrib":    rng.uniform(0, 10, n),
        "risk_contrib":        rng.uniform(0, 10, n),
        "revisions_contrib":   rng.uniform(0, 10, n),
        "size_contrib":        rng.uniform(0, 10, n),
        "investment_contrib":  rng.uniform(0, 10, n),
        "Value_Trap_Flag":     [False] * n,
        "Value_Trap_Severity": rng.uniform(0, 50, n),
        "Growth_Trap_Flag":    [False] * n,
        "Growth_Trap_Severity": rng.uniform(0, 50, n),
        "Financial_Sector_Caveat": [False] * n,
        "_is_bank_like":    [False] * n,
        "_beneish_flag":    [False] * n,
        "ev_ebitda":        rng.uniform(5, 30, n),
        "roic":             rng.uniform(0.05, 0.3, n),
        "revenue_growth":   rng.uniform(-0.05, 0.25, n),
        "Composite_Confidence": rng.uniform(25, 50, n),
    })
    return df


def _make_portfolio(n: int = 10, seed: int = 7) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    sectors = ["Information Technology", "Financials", "Health Care",
               "Consumer Discretionary", "Industrials"]
    return pd.DataFrame({
        "Port_Rank":          list(range(1, n + 1)),
        "Ticker":             [f"T{i:03d}" for i in range(n)],
        "Company":            [f"Company {i}" for i in range(n)],
        "Sector":             [sectors[i % len(sectors)] for i in range(n)],
        "Composite":          rng.uniform(60, 95, n),
        "Implied_EW_Weight":  [round(100.0 / n, 4)] * n,
        "Equal_Weight_Pct":   [round(100.0 / n, 2)] * n,
        "InvVol_Weight_Pct":  rng.uniform(5.0, 15.0, n),
        "Score_Weight_Pct":   rng.uniform(5.0, 15.0, n),
        "volatility":         rng.uniform(0.1, 0.4, n),
    })


def _make_stats(port: pd.DataFrame) -> dict:
    return {
        "n_stocks": len(port),
        "date_generated": "2026-03-16",
        "avg_composite": float(port["Composite"].mean()),
        "avg_beta": 1.05,
        "est_div_yield": 1.8,
        "_weighting_scheme": "equal",
        "sector_alloc": {
            sec: {"count": cnt, "weight_pct": cnt / len(port) * 100}
            for sec, cnt in port["Sector"].value_counts().items()
        },
        "factor_exposure": {
            "valuation": 22.0, "quality": 22.0, "growth": 13.0, "momentum": 13.0,
            "risk": 10.0, "revisions": 10.0, "size": 5.0, "investment": 5.0,
        },
    }


# ---------------------------------------------------------------------------
# write_factor_scores_sheet tests
# ---------------------------------------------------------------------------

def test_factor_scores_sheet_name():
    wb = Workbook()
    write_factor_scores_sheet(wb, _make_scored_universe())
    assert "FactorScores" in wb.sheetnames


def test_factor_scores_required_headers():
    wb = Workbook()
    write_factor_scores_sheet(wb, _make_scored_universe())
    ws = wb["FactorScores"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    required = {"Ticker", "Company", "Sector", "Composite", "Rank",
                "Val_Pct", "Qual_Pct", "Grow_Pct", "Mom_Pct",
                "Risk_Pct", "Rev_Pct", "Size_Pct", "Invest_Pct"}
    missing = required - set(headers)
    assert not missing, f"Missing headers in FactorScores: {missing}"


def test_factor_scores_row_count():
    df = _make_scored_universe(n=15)
    wb = Workbook()
    write_factor_scores_sheet(wb, df)
    ws = wb["FactorScores"]
    # Header row + n data rows (legend rows are below)
    data_rows = sum(1 for r in range(2, ws.max_row + 1)
                    if ws.cell(row=r, column=1).value not in (None, ""))
    assert data_rows >= len(df), (
        f"Expected at least {len(df)} data rows, got {data_rows}")


def test_factor_scores_color_legend_present():
    """Color legend should appear a few rows below the last data row."""
    df = _make_scored_universe(n=10)
    wb = Workbook()
    write_factor_scores_sheet(wb, df)
    ws = wb["FactorScores"]
    # Legend header should appear somewhere after row n+1
    legend_text_found = False
    for r in range(ws.max_row, 0, -1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and "Legend" in val:
            legend_text_found = True
            break
    assert legend_text_found, "Color legend header not found in FactorScores sheet"


def test_factor_scores_composite_values_numeric():
    """Composite column values should be numeric (float), not strings."""
    df = _make_scored_universe(n=10)
    wb = Workbook()
    write_factor_scores_sheet(wb, df)
    ws = wb["FactorScores"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    comp_col = headers.index("Composite") + 1
    values = [ws.cell(row=r, column=comp_col).value
              for r in range(2, len(df) + 2)]
    assert all(isinstance(v, (int, float)) for v in values if v is not None), (
        "Composite values should be numeric")


# ---------------------------------------------------------------------------
# write_full_excel integration test (sheet names)
# ---------------------------------------------------------------------------

def test_full_excel_sheet_names(monkeypatch, tmp_path):
    """write_full_excel creates all expected sheets."""
    # Point output to tmp_path
    monkeypatch.setattr(portfolio_constructor, "ROOT", tmp_path)
    cfg = {
        "output": {"excel_file": "test_output.xlsx", "factor_scores_sheet": "FactorScores"},
        "portfolio": {"num_stocks": 10},
    }

    df = _make_scored_universe(n=20)
    port = _make_portfolio(n=10)
    stats = _make_stats(port)

    out_path = write_full_excel(df, port, stats, cfg)
    wb = load_workbook(out_path)

    expected_sheets = {"FactorScores", "ScreenerDashboard", "ModelPortfolio", "DataValidation"}
    actual_sheets = set(wb.sheetnames)
    missing = expected_sheets - actual_sheets
    assert not missing, f"Missing sheets in workbook: {missing}"


def test_full_excel_returns_valid_path(monkeypatch, tmp_path):
    monkeypatch.setattr(portfolio_constructor, "ROOT", tmp_path)
    cfg = {
        "output": {"excel_file": "test_output.xlsx", "factor_scores_sheet": "FactorScores"},
        "portfolio": {"num_stocks": 10},
    }
    df = _make_scored_universe(n=20)
    port = _make_portfolio(n=10)
    stats = _make_stats(port)

    out_path = write_full_excel(df, port, stats, cfg)
    assert Path(out_path).exists()
    assert Path(out_path).suffix == ".xlsx"
