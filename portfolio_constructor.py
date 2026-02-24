#!/usr/bin/env python3
"""
Multi-Factor Stock Screener — Phase 3: Portfolio Construction & Excel Dashboard
================================================================================
Builds a model portfolio from Phase 1 factor scores and writes a 3-sheet
Excel workbook (FactorScores, ScreenerDashboard, ModelPortfolio).

Reference: Multi-Factor-Screener-Blueprint.md §6, §7.5
"""

import time
import warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

from factor_engine import (
    load_config,
    get_sp500_tickers,
    METRIC_COLS,
    CAT_METRICS,
    ROOT,
    CACHE_DIR,
    _find_latest_cache,
)

# Approximate S&P 500 sector weights (%)
# Source: S&P Dow Jones Indices, as of ~Q1 2025. These are approximate
# and should be updated periodically. Used for sector concentration caps
# in portfolio construction — NOT for scoring/ranking.
SPX_SECTOR_WEIGHTS = {
    "Information Technology": 29.0,
    "Financials": 14.0,
    "Health Care": 12.0,
    "Consumer Discretionary": 10.0,
    "Industrials": 9.0,
    "Communication Services": 9.0,
    "Consumer Staples": 6.0,
    "Energy": 4.0,
    "Utilities": 3.0,
    "Real Estate": 2.0,
    "Materials": 2.0,
}  # Sum = 100%

# Category score column names as produced by factor_engine
CATEGORY_SCORE_COLS = {
    "Valuation": "valuation_score",
    "Quality": "quality_score",
    "Growth": "growth_score",
    "Momentum": "momentum_score",
    "Risk": "risk_score",
    "Revisions": "revisions_score",
    "Size": "size_score",
    "Investment": "investment_score",
}


# =========================================================================
# A. Load latest factor scores and config
# =========================================================================
def load_latest_scores():
    """Load the most recent factor_scores parquet from cache/."""
    scores_path, _ = _find_latest_cache("factor_scores")
    if scores_path is None:
        raise FileNotFoundError(
            "No factor_scores cache found. Run factor_engine.py first."
        )
    df = pd.read_parquet(str(scores_path))
    return df, scores_path.name


# =========================================================================
# B. Construct model portfolio (§6.1)
# =========================================================================
def construct_portfolio(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    """Build model portfolio following Blueprint §6.1 steps 1-5."""
    pcfg = cfg.get("portfolio", {})
    num_stocks = pcfg.get("num_stocks", 25)
    max_sector = pcfg.get("max_sector_concentration", 8)
    max_pos_pct = pcfg.get("max_position_pct", 5.0)

    work = df.copy()

    # --- Step 0: Exclude value-trap-flagged stocks (unless flag_only) ---
    vtf = cfg.get("value_trap_filters", {})
    flag_only = vtf.get("flag_only", True)
    if not flag_only and "Value_Trap_Flag" in work.columns:
        work = work[~work["Value_Trap_Flag"]].copy()

    # --- Step 0b: Exclude growth-trap-flagged stocks (unless flag_only) ---
    gtf = cfg.get("growth_trap_filters", {})
    if not gtf.get("flag_only", True) and "Growth_Trap_Flag" in work.columns:
        work = work[~work["Growth_Trap_Flag"]].copy()

    # --- Step 1: Universe filter ---
    # Exclude stocks with Composite < 50th percentile
    median_composite = work["Composite"].median()
    work = work[work["Composite"] >= median_composite].copy()

    # Exclude stocks missing > 40% of metrics
    present = [c for c in METRIC_COLS if c in work.columns]
    work["_metric_coverage"] = work[present].notna().sum(axis=1) / len(present)
    work = work[work["_metric_coverage"] >= 0.60].copy()

    # --- Step 1b: Liquidity filter ---
    min_adv = float(pcfg.get("min_avg_dollar_volume", 10e6))
    if "avg_daily_dollar_volume" in work.columns and min_adv > 0:
        illiquid = (work["avg_daily_dollar_volume"].lt(min_adv)
                    | work["avg_daily_dollar_volume"].isna())
        work = work[~illiquid].copy()

    # --- Step 2: Rank & select top N ---
    work = work.sort_values("Composite", ascending=False).reset_index(drop=True)

    # --- Step 3: Sector constraints ---
    selected = []
    sector_counts = {}
    backfill_pool = []

    for _, row in work.iterrows():
        sec = row["Sector"]
        cnt = sector_counts.get(sec, 0)
        if cnt >= max_sector:
            backfill_pool.append(row)
            continue
        if len(selected) >= num_stocks:
            break
        selected.append(row)
        sector_counts[sec] = cnt + 1

    # Backfill from underweight sectors if needed
    if len(selected) < num_stocks:
        for _, row in work.iterrows():
            if len(selected) >= num_stocks:
                break
            ticker = row["Ticker"]
            if any(r["Ticker"] == ticker for r in selected):
                continue
            sec = row["Sector"]
            cnt = sector_counts.get(sec, 0)
            if cnt >= max_sector:
                continue
            selected.append(row)
            sector_counts[sec] = cnt + 1

    # --- Step 5: Min 20 positions guardrail ---
    if len(selected) < 20:
        # Relax composite threshold — add next-best from full universe
        for _, row in df.sort_values("Composite", ascending=False).iterrows():
            if len(selected) >= 20:
                break
            ticker = row["Ticker"]
            if any(r["Ticker"] == ticker for r in selected):
                continue
            sec = row["Sector"]
            cnt = sector_counts.get(sec, 0)
            if cnt >= max_sector:
                continue
            selected.append(row)
            sector_counts[sec] = cnt + 1

    port = pd.DataFrame(selected).reset_index(drop=True)

    if port.empty:
        port["Equal_Weight_Pct"] = []
        port["RiskParity_Weight_Pct"] = []
        port["Port_Rank"] = []
        return port

    # --- Step 4: Position sizing ---
    n = len(port)
    # Equal weight
    port["Equal_Weight_Pct"] = round(100.0 / n, 2) if n > 0 else 0

    # "Risk parity" weight: inverse-volatility (1/vol_i) normalized.
    # NOTE: This is NOT true risk parity (Bridgewater/Dalio style), which
    # requires a full covariance matrix. This is inverse-volatility weighting,
    # a simpler heuristic that tilts toward lower-volatility holdings.
    vol_col = "volatility"
    if vol_col in port.columns and port[vol_col].notna().any():
        med_vol = port[vol_col].median()
        # Guard: if median is 0 or NaN, fall back to a sensible default (25% annualized)
        if pd.isna(med_vol) or med_vol <= 0:
            med_vol = 0.25
        # Floor volatility at 5% (or 10% of median) to prevent extreme weights
        vol_floor = max(med_vol * 0.1, 0.05)
        vols = port[vol_col].fillna(med_vol)
        vols = vols.replace(0, med_vol)
        vols = vols.clip(lower=vol_floor)
        inv_vol = 1.0 / vols
        total_iv = inv_vol.sum()
        if total_iv > 0 and np.isfinite(total_iv):
            port["RiskParity_Weight_Pct"] = round(inv_vol / total_iv * 100, 2)
        else:
            port["RiskParity_Weight_Pct"] = port["Equal_Weight_Pct"]
    else:
        port["RiskParity_Weight_Pct"] = port["Equal_Weight_Pct"]

    # --- Step 5: Max single position cap and redistribute ---
    cap = max_pos_pct  # Use config value (default 5.0%)
    for wt_col in ["Equal_Weight_Pct", "RiskParity_Weight_Pct"]:
        # Convert to float to avoid type issues
        port[wt_col] = pd.to_numeric(port[wt_col], errors='coerce').fillna(0.0)
        # Iterative cap-and-redistribute until no position exceeds cap
        for _iteration in range(10):  # Max 10 iterations to converge
            excess = 0.0
            uncapped_idx = []
            for i in port.index:
                wt_val = port.loc[i, wt_col]
                if wt_val > cap:  # type: ignore
                    excess += wt_val - cap  # type: ignore
                    port.loc[i, wt_col] = cap
                else:
                    uncapped_idx.append(i)
            if excess <= 0.001 or len(uncapped_idx) == 0:
                break
            add_per = excess / len(uncapped_idx)
            for i in uncapped_idx:
                port.loc[i, wt_col] = port.loc[i, wt_col] + add_per  # type: ignore
        # Normalize to 100%
        total = port[wt_col].sum()
        if total > 0:
            port[wt_col] = round(port[wt_col] / total * 100, 2)

    # Rank within portfolio
    port = port.sort_values("Composite", ascending=False).reset_index(drop=True)
    port["Port_Rank"] = range(1, len(port) + 1)

    return port


# =========================================================================
# C. Portfolio summary stats (§6.3)
# =========================================================================
def compute_portfolio_stats(port: pd.DataFrame, cfg: dict) -> dict:
    """Compute summary statistics for the model portfolio."""
    n = len(port)
    if n == 0:
        return {
            "n_stocks": 0, "avg_composite": 0, "avg_beta": 1.0,
            "est_div_yield": 0.0, "sector_alloc": {},
            "factor_exposure": {}, "date_generated": datetime.now().strftime("%Y-%m-%d"),
        }
    # Use configured weighting scheme
    scheme = cfg.get("portfolio", {}).get("weighting", "equal")
    if scheme == "risk_parity" and "RiskParity_Weight_Pct" in port.columns:
        wt_col = "RiskParity_Weight_Pct"
    else:
        wt_col = "Equal_Weight_Pct"
    weights_array = pd.to_numeric(port[wt_col], errors='coerce').to_numpy(dtype=np.float64)
    weights = weights_array / 100.0

    # Weighted avg composite
    composite_array = pd.to_numeric(port["Composite"], errors='coerce').to_numpy(dtype=np.float64)
    avg_composite = np.average(composite_array, weights=weights)

    # Weighted avg beta
    beta_col = "beta"
    if beta_col in port.columns:
        betas = pd.to_numeric(port[beta_col].fillna(1.0), errors='coerce').to_numpy(dtype=np.float64)
        avg_beta = np.average(betas, weights=weights)
    else:
        avg_beta = 1.0

    # Dividend yield (default 0 if not available)
    # yfinance dividendYield is already in percent (e.g. 0.41 = 0.41%)
    div_col = "dividend_yield"
    if div_col in port.columns:
        divs = pd.to_numeric(port[div_col].fillna(0), errors='coerce').to_numpy(dtype=np.float64)
        avg_div = np.average(divs, weights=weights)  # already in %
    else:
        avg_div = 0.0

    # Sector allocation
    sector_alloc = {}
    for sec in port["Sector"].unique():
        mask = port["Sector"] == sec
        sector_vals = pd.to_numeric(port.loc[mask, wt_col], errors='coerce')
        weight_sum = sector_vals.sum() if isinstance(sector_vals, pd.Series) else sector_vals
        sector_alloc[sec] = {
            "count": int(mask.sum()),
            "weight_pct": round(weight_sum, 2),
        }

    # Factor exposure: weighted average of each category percentile
    factor_exposure = {}
    for label, col in CATEGORY_SCORE_COLS.items():
        if col in port.columns:
            vals = pd.to_numeric(port[col].fillna(50), errors='coerce').to_numpy(dtype=np.float64)
            factor_exposure[label] = round(np.average(vals, weights=weights), 1)
        else:
            factor_exposure[label] = 0.0

    return {
        "n_stocks": n,
        "avg_composite": round(avg_composite, 1),
        "avg_beta": round(avg_beta, 2),
        "est_div_yield": round(avg_div, 2),
        "sector_alloc": sector_alloc,
        "factor_exposure": factor_exposure,
        "date_generated": datetime.now().strftime("%Y-%m-%d"),
    }


# =========================================================================
# Excel writing — 3 sheets
# =========================================================================

# Styling constants
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def _style_header_row(ws, row, n_cols):
    """Apply header styling to a row."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_cell(ws, row, col, border=True):
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if border:
        cell.border = THIN_BORDER
    return cell


def _auto_width(ws, min_width=8, max_width=25):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


# ---- Sheet 1: FactorScores (unchanged from Phase 1) ----
def write_factor_scores_sheet(wb: Workbook, df: pd.DataFrame):
    """Write FactorScores sheet — same as Phase 1."""
    if wb.active is None:
        ws = wb.create_sheet("FactorScores", 0)
    else:
        ws = wb.active
        ws.title = "FactorScores"

    col_map = [
        ("Ticker", "Ticker"), ("Company", "Company"), ("Sector", "Sector"),
        ("valuation_score", "Val_Pct"), ("quality_score", "Qual_Pct"),
        ("growth_score", "Grow_Pct"), ("momentum_score", "Mom_Pct"),
        ("risk_score", "Risk_Pct"), ("revisions_score", "Rev_Pct"),
        ("size_score", "Size_Pct"), ("investment_score", "Invest_Pct"),
        ("Composite", "Composite"), ("Rank", "Rank"),
        ("valuation_contrib", "Val_Contrib"), ("quality_contrib", "Qual_Contrib"),
        ("growth_contrib", "Grow_Contrib"), ("momentum_contrib", "Mom_Contrib"),
        ("risk_contrib", "Risk_Contrib"), ("revisions_contrib", "Rev_Contrib"),
        ("size_contrib", "Size_Contrib"), ("investment_contrib", "Invest_Contrib"),
        ("Value_Trap_Flag", "Value_Trap_Flag"),
        ("Growth_Trap_Flag", "Growth_Trap_Flag"),
        ("Financial_Sector_Caveat", "Fin_Caveat"),
        ("_is_bank_like", "Is_Bank"),
    ]

    # Contribution columns use a distinct fill to visually separate them
    contrib_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    # Header
    for c, (_, header) in enumerate(col_map, 1):
        ws.cell(row=1, column=c, value=header)
    _style_header_row(ws, 1, len(col_map))

    # Data
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, (src, _) in enumerate(col_map, 1):
            v = row.get(src)
            if isinstance(v, float) and np.isnan(v):
                v = None
            elif src in ("valuation_score", "quality_score", "growth_score",
                         "momentum_score", "risk_score", "revisions_score",
                         "size_score", "investment_score"):
                v = round(v, 1) if pd.notna(v) else None
            elif src.endswith("_contrib"):
                v = round(v, 2) if pd.notna(v) else None
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if c > 3 else "left")
            if src.endswith("_contrib"):
                cell.fill = contrib_fill

    _auto_width(ws)
    return ws


# ---- Sheet 2: ScreenerDashboard ----
def write_screener_dashboard(wb: Workbook, df: pd.DataFrame):
    """Write ScreenerDashboard with Top 25, Sector Summary, Score Distribution."""
    ws = wb.create_sheet("ScreenerDashboard")

    # ---- Title ----
    ws.cell(row=1, column=1, value="SCREENER DASHBOARD").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d')}").font = Font(
        name="Calibri", size=10, italic=True, color="666666")

    # ==== TOP 25 TABLE ====
    row = 4
    ws.cell(row=row, column=1, value="TOP 25 STOCKS BY COMPOSITE SCORE").font = SUBTITLE_FONT
    row += 1
    top25_headers = ["Rank", "Ticker", "Company", "Sector", "Composite",
                     "Val_Pct", "Qual_Pct", "Mom_Pct", "Value_Trap_Flag"]
    for c, h in enumerate(top25_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(top25_headers))
    row += 1

    top25 = df.nsmallest(25, "Rank")
    for _, stock in top25.iterrows():
        ws.cell(row=row, column=1, value=int(stock["Rank"]))
        ws.cell(row=row, column=2, value=stock["Ticker"])
        ws.cell(row=row, column=3, value=stock.get("Company", ""))
        ws.cell(row=row, column=4, value=stock.get("Sector", ""))
        comp = round(stock["Composite"], 1) if pd.notna(stock["Composite"]) else None
        ws.cell(row=row, column=5, value=comp)
        ws.cell(row=row, column=6, value=round(stock.get("valuation_score", 0), 1))
        ws.cell(row=row, column=7, value=round(stock.get("quality_score", 0), 1))
        ws.cell(row=row, column=8, value=round(stock.get("momentum_score", 0), 1))
        ws.cell(row=row, column=9, value=bool(stock.get("Value_Trap_Flag", False)))

        # Conditional formatting on Composite
        for c in range(1, len(top25_headers) + 1):
            _style_data_cell(ws, row, c)
        if comp is not None:
            if comp >= 80:
                ws.cell(row=row, column=5).fill = GREEN_FILL
            elif comp >= 60:
                ws.cell(row=row, column=5).fill = YELLOW_FILL
            else:
                ws.cell(row=row, column=5).fill = RED_FILL

        row += 1

    # ==== SECTOR SUMMARY TABLE ====
    row += 2
    ws.cell(row=row, column=1, value="SECTOR SUMMARY").font = SUBTITLE_FONT
    row += 1
    sec_headers = ["Sector", "Avg Composite", "# Stocks", "# in Top 25"]
    for c, h in enumerate(sec_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(sec_headers))
    row += 1

    top25_tickers = set(top25["Ticker"].tolist())
    sector_groups = df.groupby("Sector")
    for sec_name in sorted(df["Sector"].unique()):
        grp = sector_groups.get_group(sec_name)
        avg_comp = round(grp["Composite"].mean(), 1)
        n_stocks = len(grp)
        n_top25 = sum(1 for t in grp["Ticker"] if t in top25_tickers)

        ws.cell(row=row, column=1, value=sec_name)
        ws.cell(row=row, column=2, value=avg_comp)
        ws.cell(row=row, column=3, value=n_stocks)
        ws.cell(row=row, column=4, value=n_top25)

        for c in range(1, len(sec_headers) + 1):
            _style_data_cell(ws, row, c)

        # 3-color conditional format on Avg Composite
        if avg_comp >= 60:
            ws.cell(row=row, column=2).fill = GREEN_FILL
        elif avg_comp >= 45:
            ws.cell(row=row, column=2).fill = YELLOW_FILL
        else:
            ws.cell(row=row, column=2).fill = RED_FILL

        row += 1

    # ==== SCORE DISTRIBUTION TABLE ====
    row += 2
    ws.cell(row=row, column=1, value="SCORE DISTRIBUTION").font = SUBTITLE_FONT
    row += 1
    dist_headers = ["Score Range", "# Stocks"]
    for c, h in enumerate(dist_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(dist_headers))
    row += 1

    buckets = [(i, i + 9) for i in range(0, 100, 10)]
    for low, high in buckets:
        label = f"{low}–{high}" if high < 99 else f"{low}–100"
        if high >= 99:
            # Last bucket: include 100 (>= low, no upper bound)
            count = int((df["Composite"] >= low).sum())
        else:
            count = int(((df["Composite"] >= low) & (df["Composite"] < low + 10)).sum())
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        for c in range(1, 3):
            _style_data_cell(ws, row, c)
        row += 1

    _auto_width(ws)
    return ws


# ---- Sheet 3: ModelPortfolio ----
def write_model_portfolio_sheet(wb: Workbook, port: pd.DataFrame, stats: dict):
    """Write ModelPortfolio sheet with holdings, sector alloc, factor exposure."""
    ws = wb.create_sheet("ModelPortfolio")

    # ---- Title ----
    ws.cell(row=1, column=1, value="MODEL PORTFOLIO").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {stats['date_generated']}").font = Font(
        name="Calibri", size=10, italic=True, color="666666")

    # ==== HOLDINGS TABLE ====
    row = 4
    ws.cell(row=row, column=1, value="PORTFOLIO HOLDINGS").font = SUBTITLE_FONT
    row += 1
    hold_headers = ["Rank", "Ticker", "Company", "Sector", "Composite",
                    "Equal_Weight_%", "RiskParity_Weight_%"]
    for c, h in enumerate(hold_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(hold_headers))
    row += 1

    for _, stock in port.iterrows():
        ws.cell(row=row, column=1, value=int(stock["Port_Rank"]))
        ws.cell(row=row, column=2, value=stock["Ticker"])
        ws.cell(row=row, column=3, value=stock.get("Company", ""))
        ws.cell(row=row, column=4, value=stock.get("Sector", ""))
        ws.cell(row=row, column=5, value=round(stock["Composite"], 1))
        ws.cell(row=row, column=6, value=round(stock["Equal_Weight_Pct"], 2))
        ws.cell(row=row, column=7, value=round(stock["RiskParity_Weight_Pct"], 2))

        for c in range(1, len(hold_headers) + 1):
            _style_data_cell(ws, row, c)

        # Composite coloring
        comp = stock["Composite"]
        if comp >= 80:
            ws.cell(row=row, column=5).fill = GREEN_FILL
        elif comp >= 60:
            ws.cell(row=row, column=5).fill = YELLOW_FILL
        else:
            ws.cell(row=row, column=5).fill = RED_FILL

        row += 1

    # ==== SECTOR ALLOCATION TABLE ====
    row += 2
    ws.cell(row=row, column=1, value="SECTOR ALLOCATION").font = SUBTITLE_FONT
    row += 1
    sec_headers = ["Sector", "# Holdings", "Portfolio Weight %",
                   "SPX Approx Weight %", "Delta"]
    for c, h in enumerate(sec_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(sec_headers))
    row += 1

    all_sectors = sorted(set(list(stats["sector_alloc"].keys()) +
                             list(SPX_SECTOR_WEIGHTS.keys())))
    for sec in all_sectors:
        alloc = stats["sector_alloc"].get(sec, {"count": 0, "weight_pct": 0.0})
        spx_wt = SPX_SECTOR_WEIGHTS.get(sec, 0.0)
        delta = round(alloc["weight_pct"] - spx_wt, 1)

        ws.cell(row=row, column=1, value=sec)
        ws.cell(row=row, column=2, value=alloc["count"])
        ws.cell(row=row, column=3, value=round(alloc["weight_pct"], 1))
        ws.cell(row=row, column=4, value=spx_wt)
        ws.cell(row=row, column=5, value=delta)

        for c in range(1, len(sec_headers) + 1):
            _style_data_cell(ws, row, c)

        # Color delta: green within ±5%, yellow ±5-10%, red >±10%
        abs_delta = abs(delta)
        if abs_delta <= 5:
            ws.cell(row=row, column=5).fill = GREEN_FILL
        elif abs_delta <= 10:
            ws.cell(row=row, column=5).fill = YELLOW_FILL
        else:
            ws.cell(row=row, column=5).fill = RED_FILL

        row += 1

    # ==== PORTFOLIO SUMMARY BLOCK ====
    row += 2
    ws.cell(row=row, column=1, value="PORTFOLIO SUMMARY").font = SUBTITLE_FONT
    row += 1

    summary_data = [
        ("Total Stocks", stats["n_stocks"]),
        ("Avg Composite Score", stats["avg_composite"]),
        ("Portfolio Beta", stats["avg_beta"]),
        ("Est. Dividend Yield (%)", stats["est_div_yield"]),
        ("Date Generated", stats["date_generated"]),
    ]
    for label, val in summary_data:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=row, column=2, value=val).font = DATA_FONT
        ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # ==== FACTOR EXPOSURE TABLE ====
    row += 2
    ws.cell(row=row, column=1, value="FACTOR EXPOSURE (Portfolio Avg Percentile)").font = SUBTITLE_FONT
    row += 1
    fe_headers = ["Factor Category", "Avg Percentile"]
    for c, h in enumerate(fe_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(fe_headers))
    row += 1

    for cat, val in stats["factor_exposure"].items():
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=val)
        for c in range(1, 3):
            _style_data_cell(ws, row, c)

        # Color by target ranges from §6.3
        if val >= 65:
            ws.cell(row=row, column=2).fill = GREEN_FILL
        elif val >= 50:
            ws.cell(row=row, column=2).fill = YELLOW_FILL
        else:
            ws.cell(row=row, column=2).fill = RED_FILL

        row += 1

    _auto_width(ws)
    return ws


# ---- Sheet 4: WeightSensitivity ----
def write_weight_sensitivity_sheet(wb: Workbook, sens_df: pd.DataFrame):
    """Write weight sensitivity analysis results."""
    if sens_df is None or sens_df.empty:
        return
    ws = wb.create_sheet("WeightSensitivity")

    headers = ["Category", "Direction", "Orig Weight", "Perturbed Weight",
               "Top-20 Unchanged", "Top-20 Changed", "Jaccard Similarity",
               "Changed Tickers"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    # Summary interpretation row
    ws.cell(row=2, column=1, value="INTERPRETATION:")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="666666")
    ws.cell(row=2, column=1).value = (
        "Jaccard ≥ 0.85 = robust (top-20 is stable). "
        "Jaccard < 0.70 = sensitive (weight choice materially affects picks). "
        "Perturbation = ±5% of category weight, others scaled proportionally.")

    src_cols = ["category", "direction", "original_weight", "perturbed_weight",
                "top_n_unchanged", "top_n_changed", "jaccard_similarity",
                "changed_tickers"]
    for r, (_, row) in enumerate(sens_df.iterrows(), 3):
        for c, col in enumerate(src_cols, 1):
            v = row.get(col)
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if c < 8 else "left")
            # Color-code Jaccard
            if col == "jaccard_similarity" and pd.notna(v):
                if v >= 0.85:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")  # green
                elif v < 0.70:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")  # red
                else:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")  # yellow

    _auto_width(ws)


# ---- Sheet 5: FactorCorrelation ----
def write_factor_correlation_sheet(wb: Workbook, corr_df: pd.DataFrame):
    """Write cross-metric Spearman correlation matrix."""
    if corr_df is None or corr_df.empty:
        return
    ws = wb.create_sheet("FactorCorrelation")

    # Clean up column names for display (remove _pct suffix)
    display_names = [c.replace("_pct", "") for c in corr_df.columns]

    # Header row
    ws.cell(row=1, column=1, value="")
    for c, name in enumerate(display_names, 2):
        ws.cell(row=1, column=c, value=name)
    _style_header_row(ws, 1, len(display_names) + 1)

    # Row labels + data
    for r, (idx, row_data) in enumerate(corr_df.iterrows(), 2):
        label = str(idx).replace("_pct", "")
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = Font(bold=True, size=9)
        ws.cell(row=r, column=1).border = THIN_BORDER
        for c, val in enumerate(row_data, 2):
            cell = ws.cell(row=r, column=c)
            if pd.notna(val):
                cell.value = round(float(val), 3)
                # Color-code: high correlation (>0.6) = orange, very high (>0.8) = red
                abs_val = abs(float(val))
                if abs_val > 0.99:  # diagonal
                    cell.fill = PatternFill("solid", fgColor="D9D9D9")
                elif abs_val > 0.8:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                elif abs_val > 0.6:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    _auto_width(ws, min_width=6, max_width=12)

    # Add interpretation note
    note_row = len(corr_df) + 3
    ws.cell(row=note_row, column=1, value="INTERPRETATION:")
    ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=min(8, len(display_names) + 1))
    ws.cell(row=note_row, column=1).value = (
        "Spearman rank correlation between sector-relative percentile scores. "
        "Red (>0.8) = high collinearity — these metrics measure nearly the same thing. "
        "Orange (>0.6) = moderate overlap — intentional overlap is acceptable if documented.")


# ---- Sheet 6: DataValidation (top-N with raw values + provenance) ----
def write_data_validation_sheet(wb: Workbook, df: pd.DataFrame, top_n: int = 10):
    """Write a validation sheet showing raw data for top-ranked stocks.

    Purpose: allow the user to spot-check key values against Bloomberg/CapIQ
    before acting on the screener output.
    """
    ws = wb.create_sheet("DataValidation")

    # Title
    ws.cell(row=1, column=1, value=f"TOP-{top_n} STOCKS — RAW DATA FOR MANUAL VERIFICATION")
    ws.cell(row=1, column=1).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    ws.cell(row=2, column=1, value=(
        "Verify these values against Bloomberg, FactSet, or SEC filings before "
        "acting on screener output. Flag any value that differs materially."))
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    # Select top-N
    top = df.nsmallest(top_n, "Rank") if "Rank" in df.columns else df.head(top_n)

    # Columns to show for validation (raw values the user should verify)
    val_cols = [
        ("Ticker", "Ticker"),
        ("Rank", "Rank"),
        ("Composite", "Composite"),
        ("_data_source", "Data Source"),
        ("_metric_count", "Metrics Available"),
        ("_metric_total", "Metrics Total"),
        ("_eps_basis_mismatch", "EPS Mismatch?"),
        ("_eps_ratio", "Fwd/Trail EPS Ratio"),
        ("ev_ebitda", "EV/EBITDA (raw)"),
        ("fcf_yield", "FCF Yield (raw)"),
        ("earnings_yield", "Earnings Yield (raw)"),
        ("roic", "ROIC (raw)"),
        ("debt_equity", "D/E (raw)"),
        ("forward_eps_growth", "Fwd EPS Growth (raw)"),
        ("revenue_growth", "Rev Growth (raw)"),
        ("return_12_1", "12-1M Return (raw)"),
        ("volatility", "Volatility (raw)"),
        ("beta", "Beta (raw)"),
        ("analyst_surprise", "Analyst Surprise (raw)"),
        ("_current_price", "Price"),
        ("_target_mean", "Analyst Target"),
        ("_num_analysts", "# Analysts"),
        ("_stale_data", "Stale Data?"),
        ("_ev_flag", "EV Discrepancy"),
    ]

    # Headers
    for c, (_, header) in enumerate(val_cols, 1):
        ws.cell(row=3, column=c, value=header)
    _style_header_row(ws, 3, len(val_cols))

    # Data
    for r, (_, row) in enumerate(top.iterrows(), 4):
        for c, (src, _) in enumerate(val_cols, 1):
            v = row.get(src)
            if isinstance(v, float) and np.isnan(v):
                v = None
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
            # Highlight EPS mismatch
            if src == "_eps_basis_mismatch" and v is True:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")  # yellow warning
            # Highlight stale data
            if src == "_stale_data" and v is True:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")  # red warning
            # Highlight EV discrepancy
            if src == "_ev_flag" and v is not None and v is not False:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
            # Format percentages
            if src in ("fcf_yield", "earnings_yield", "roic", "forward_eps_growth",
                       "revenue_growth", "return_12_1", "analyst_surprise"):
                if v is not None:
                    cell.number_format = '0.00%'
            if src == "volatility" and v is not None:
                cell.number_format = '0.00%'

    _auto_width(ws)


# =========================================================================
# Full Excel writer — 6 sheets
# =========================================================================
def write_full_excel(df: pd.DataFrame, port: pd.DataFrame,
                     stats: dict, cfg: dict,
                     sens_df: pd.DataFrame = None,
                     corr_df: pd.DataFrame = None) -> str:
    """Write factor_output.xlsx with up to 6 sheets.

    Raises PermissionError if file is locked (handled by run_screener.py).
    """
    out_path = ROOT / cfg["output"]["excel_file"]
    wb = Workbook()

    # Sheet 1: FactorScores
    write_factor_scores_sheet(wb, df)

    # Sheet 2: ScreenerDashboard
    write_screener_dashboard(wb, df)

    # Sheet 3: ModelPortfolio
    write_model_portfolio_sheet(wb, port, stats)

    # Sheet 4: WeightSensitivity (if available)
    if sens_df is not None and not sens_df.empty:
        write_weight_sensitivity_sheet(wb, sens_df)

    # Sheet 5: FactorCorrelation (if available)
    if corr_df is not None and not corr_df.empty:
        write_factor_correlation_sheet(wb, corr_df)

    # Sheet 6: DataValidation (always — top-10 raw values for spot-checking)
    write_data_validation_sheet(wb, df, top_n=10)

    try:
        wb.save(str(out_path))
    except PermissionError:
        raise PermissionError(
            f"Cannot write '{out_path.name}'. Close the file in Excel and re-run.")
    except Exception as e:
        raise RuntimeError(f"Failed to write Excel: {e}")
    return str(out_path)


# =========================================================================
# Diagnostics
# =========================================================================
def print_summary(df, port, stats, capped_sectors, cfg, excel_path, t0):
    elapsed = round(time.time() - t0, 1)

    print()
    print("============================================")
    print("  SCREENER + PORTFOLIO — RUN SUMMARY")
    print("============================================")
    print(f"Universe scored:          {len(df)} tickers")
    print(f"Portfolio stocks:         {stats['n_stocks']} (target: {cfg.get('portfolio', {}).get('num_stocks', 25)})")

    if capped_sectors:
        cap_str = ", ".join(capped_sectors)
        print(f"Sector cap applied:       YES ({cap_str})")
    else:
        print(f"Sector cap applied:       NO")

    print("--------------------------------------------")
    print("MODEL PORTFOLIO — TOP 5:")
    top5 = port.nsmallest(5, "Port_Rank")
    for _, r in top5.iterrows():
        print(f"  {int(r['Port_Rank']):2d}. {r['Ticker']:6s} {str(r.get('Sector','')):26s} "
              f"{r['Composite']:.1f}  EqWt={r['Equal_Weight_Pct']:.1f}%  "
              f"RPWt={r['RiskParity_Weight_Pct']:.1f}%")

    print("--------------------------------------------")
    print("SECTOR ALLOCATION:")
    print(f"  {'Sector':<25s}| {'# Stk':>5s} | {'Port%':>6s} | {'SPX%':>6s} | {'Delta':>6s}")
    for sec in sorted(stats["sector_alloc"].keys()):
        alloc = stats["sector_alloc"][sec]
        spx = SPX_SECTOR_WEIGHTS.get(sec, 0.0)
        delta = alloc["weight_pct"] - spx
        print(f"  {sec:<25s}| {alloc['count']:>5d} | {alloc['weight_pct']:>5.1f}% | "
              f"{spx:>5.1f}% | {delta:>+5.1f}%")

    print("--------------------------------------------")
    print("FACTOR EXPOSURE (Portfolio Avg):")
    for cat, val in stats["factor_exposure"].items():
        print(f"  {cat + ':':<18s} {val:.1f}")

    print("--------------------------------------------")
    print(f"Portfolio Beta:           {stats['avg_beta']:.2f}")
    print(f"Est. Dividend Yield:      {stats['est_div_yield']:.2f}%")
    print("--------------------------------------------")
    print(f"Excel written:            {cfg['output']['excel_file']}")
    print(f"  Sheets: FactorScores, ScreenerDashboard, ModelPortfolio")
    print(f"Total runtime:            {elapsed}s")
    print("============================================")


# =========================================================================
# MAIN
# =========================================================================
def main():
    t0 = time.time()

    # ---- Load config ----
    print("Loading configuration...")
    cfg = load_config()

    # Handle revisions auto-disable (same logic as factor_engine)
    # Check after loading scores

    # ---- Load Phase 1 scored universe ----
    print("Loading latest factor scores...")
    df, cache_name = load_latest_scores()
    print(f"  Loaded {len(df)} tickers from {cache_name}")

    # Ensure Rank column exists
    if "Rank" not in df.columns:
        df["Rank"] = df["Composite"].rank(ascending=False, method="min").astype(int)
        df = df.sort_values("Rank").reset_index(drop=True)

    # Ensure Value_Trap_Flag exists
    if "Value_Trap_Flag" not in df.columns:
        df["Value_Trap_Flag"] = False

    # Check revisions coverage for config adjustment
    rev_m = ["analyst_surprise", "price_target_upside"]
    rev_avail = sum(df[c].notna().sum() for c in rev_m if c in df.columns)
    rev_total = len(df) * len(rev_m)
    rev_pct = rev_avail / rev_total * 100 if rev_total else 0
    if rev_pct < 30:
        old_w = cfg["factor_weights"]["revisions"]
        cfg["factor_weights"]["revisions"] = 0
        others = [k for k in cfg["factor_weights"] if k != "revisions"]
        s = sum(cfg["factor_weights"][k] for k in others)
        if s > 0:
            for k in others:
                cfg["factor_weights"][k] += old_w * cfg["factor_weights"][k] / s
            for k in cfg["factor_weights"]:
                cfg["factor_weights"][k] = round(cfg["factor_weights"][k], 2)

    # ---- Construct portfolio ----
    print("Constructing model portfolio...")
    port = construct_portfolio(df, cfg)
    print(f"  Selected {len(port)} stocks")

    # Detect which sectors were capped
    max_sector = cfg.get("portfolio", {}).get("max_sector_concentration", 8)
    sector_counts = port["Sector"].value_counts()
    capped_sectors = [sec for sec, cnt in sector_counts.items() if cnt >= max_sector]

    # ---- Compute stats ----
    print("Computing portfolio statistics...")
    stats = compute_portfolio_stats(port, cfg)

    # ---- Write 3-sheet Excel ----
    print("Writing 3-sheet Excel workbook...")
    excel_path = write_full_excel(df, port, stats, cfg)
    print(f"  Written to {excel_path}")

    # ---- Diagnostics ----
    print_summary(df, port, stats, capped_sectors, cfg, excel_path, t0)


if __name__ == "__main__":
    main()
