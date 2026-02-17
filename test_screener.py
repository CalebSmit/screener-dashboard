#!/usr/bin/env python3
"""
Multi-Factor Stock Screener — Smoke Test Suite
================================================
Run:  python -m pytest test_screener.py -v
      python test_screener.py            # also works standalone

Covers:
  - Unit tests for every core function in factor_engine.py
  - Unit tests for portfolio_constructor.py
  - Unit tests for run_screener.py utilities
  - Integration test: full pipeline end-to-end (synthetic data)
  - Edge cases: empty DF, 1-ticker, all-NaN, zero-vol, missing config keys

All tests use synthetic "fake data mode" — no network calls.
"""

import importlib
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

# ---------------------------------------------------------------------------
# Fixtures — shared across all tests
# ---------------------------------------------------------------------------

@pytest.fixture(scope="session")
def cfg():
    """Load config once for the whole test session."""
    from factor_engine import load_config
    return load_config()


@pytest.fixture(scope="session")
def universe_df(cfg):
    from factor_engine import get_sp500_tickers
    return get_sp500_tickers(cfg)


@pytest.fixture(scope="session")
def sample_df(universe_df):
    """Full 392-ticker synthetic DataFrame (raw metrics, not yet scored)."""
    from factor_engine import _generate_sample_data
    return _generate_sample_data(universe_df, seed=42)


@pytest.fixture(scope="session")
def scored_df(sample_df, cfg):
    """Fully scored DataFrame ready for portfolio construction."""
    from factor_engine import (
        winsorize_metrics, compute_sector_percentiles,
        compute_category_scores, compute_composite,
        apply_value_trap_flags, rank_stocks,
    )
    # Copy cfg to avoid mutation from revisions auto-disable
    import copy
    _cfg = copy.deepcopy(cfg)

    # Revisions auto-disable
    rev_m = ["analyst_surprise", "eps_revision_ratio", "eps_estimate_change"]
    rev_avail = sum(sample_df[c].notna().sum() for c in rev_m if c in sample_df.columns)
    rev_total = len(sample_df) * len(rev_m)
    rev_pct = rev_avail / rev_total * 100 if rev_total else 0
    if rev_pct < 30:
        old_w = _cfg["factor_weights"]["revisions"]
        _cfg["factor_weights"]["revisions"] = 0
        others = [k for k in _cfg["factor_weights"] if k != "revisions"]
        s = sum(_cfg["factor_weights"][k] for k in others)
        if s > 0:
            for k in others:
                _cfg["factor_weights"][k] += old_w * _cfg["factor_weights"][k] / s

    df = winsorize_metrics(sample_df.copy())
    df = compute_sector_percentiles(df)
    df = compute_category_scores(df, _cfg)
    df = compute_composite(df, _cfg)
    df = apply_value_trap_flags(df, _cfg)
    df = rank_stocks(df)
    return df


# ===========================================================================
# UNIT TESTS: factor_engine.py
# ===========================================================================

class TestLoadConfig:
    def test_returns_dict(self, cfg):
        assert isinstance(cfg, dict)

    def test_has_factor_weights(self, cfg):
        assert "factor_weights" in cfg
        assert sum(cfg["factor_weights"].values()) == 100

    def test_has_metric_weights(self, cfg):
        assert "metric_weights" in cfg
        for cat in ["valuation", "quality", "growth", "momentum", "risk", "revisions"]:
            assert cat in cfg["metric_weights"]


class TestGetSP500Tickers:
    def test_returns_dataframe(self, universe_df):
        assert isinstance(universe_df, pd.DataFrame)
        assert len(universe_df) > 300

    def test_has_required_columns(self, universe_df):
        for col in ["Ticker", "Company", "Sector"]:
            assert col in universe_df.columns

    def test_no_duplicate_tickers(self, universe_df):
        assert universe_df["Ticker"].is_unique


class TestGenerateSampleData:
    def test_deterministic_seed(self, universe_df):
        from factor_engine import _generate_sample_data
        df1 = _generate_sample_data(universe_df, seed=42)
        df2 = _generate_sample_data(universe_df, seed=42)
        pd.testing.assert_frame_equal(df1, df2)

    def test_shape_matches_universe(self, universe_df, sample_df):
        assert len(sample_df) == len(universe_df)

    def test_has_all_metric_cols(self, sample_df):
        from factor_engine import METRIC_COLS
        for col in METRIC_COLS:
            assert col in sample_df.columns, f"Missing column: {col}"

    def test_sectors_preserved(self, universe_df, sample_df):
        assert set(sample_df["Sector"]) == set(universe_df["Sector"])


class TestWinsorize:
    def test_does_not_change_shape(self, sample_df):
        from factor_engine import winsorize_metrics
        before = len(sample_df)
        df = winsorize_metrics(sample_df.copy())
        assert len(df) == before

    def test_reduces_outlier_range(self, sample_df):
        from factor_engine import winsorize_metrics
        col = "ev_ebitda"
        raw_range = sample_df[col].max() - sample_df[col].min()
        df = winsorize_metrics(sample_df.copy())
        win_range = df[col].max() - df[col].min()
        assert win_range <= raw_range


class TestSectorPercentiles:
    def test_pct_cols_created(self, sample_df):
        from factor_engine import compute_sector_percentiles, METRIC_COLS, winsorize_metrics
        df = winsorize_metrics(sample_df.copy())
        df = compute_sector_percentiles(df)
        for col in METRIC_COLS:
            assert f"{col}_pct" in df.columns

    def test_pct_range_0_100(self, sample_df):
        from factor_engine import compute_sector_percentiles, winsorize_metrics
        df = winsorize_metrics(sample_df.copy())
        df = compute_sector_percentiles(df)
        for col in df.columns:
            if col.endswith("_pct"):
                assert df[col].min() >= 0, f"{col} has value < 0"
                assert df[col].max() <= 100.1, f"{col} has value > 100"


class TestComposite:
    def test_range_0_100(self, scored_df):
        assert scored_df["Composite"].min() >= 0
        assert scored_df["Composite"].max() <= 100.01

    def test_no_nans(self, scored_df):
        assert scored_df["Composite"].notna().all()

    def test_empty_df(self, cfg):
        from factor_engine import compute_composite
        empty = pd.DataFrame(columns=[
            "valuation_score", "quality_score", "growth_score",
            "momentum_score", "risk_score", "revisions_score"
        ])
        result = compute_composite(empty, cfg)
        assert result.empty

    def test_zero_weights(self):
        from factor_engine import compute_composite
        cfg_zero = {"factor_weights": {k: 0 for k in
            ["valuation", "quality", "growth", "momentum", "risk", "revisions"]}}
        df = pd.DataFrame({
            "valuation_score": [60], "quality_score": [70],
            "growth_score": [55], "momentum_score": [65],
            "risk_score": [50], "revisions_score": [40],
        })
        result = compute_composite(df, cfg_zero)
        assert result["Composite"].notna().all()


class TestValueTrapFlags:
    def test_flag_column_exists(self, scored_df):
        assert "Value_Trap_Flag" in scored_df.columns

    def test_flag_is_boolean(self, scored_df):
        assert scored_df["Value_Trap_Flag"].dtype == bool

    def test_some_flagged(self, scored_df):
        # With synthetic data, some should be flagged
        assert scored_df["Value_Trap_Flag"].sum() > 0


class TestRankStocks:
    def test_rank_column(self, scored_df):
        assert "Rank" in scored_df.columns
        assert scored_df["Rank"].min() == 1
        assert scored_df["Rank"].max() == len(scored_df)

    def test_sorted_by_rank(self, scored_df):
        assert (scored_df["Rank"].diff().dropna() >= 0).all()


class TestComputeMetrics:
    def test_with_empty_market_returns(self):
        from factor_engine import compute_metrics
        raw = [{"Ticker": "X", "marketCap": 1e12, "enterpriseValue": 1e12,
                "trailingEps": 5, "currentPrice": 100, "sector": "Tech",
                "shortName": "X Corp", "totalRevenue": 50e9, "grossProfit": 20e9,
                "ebit": 15e9, "ebitda": 18e9, "netIncome": 10e9,
                "totalAssets": 200e9, "totalEquity": 50e9,
                "operatingCashFlow": 12e9, "capex": -3e9}]
        df = compute_metrics(raw, pd.Series(dtype=float))
        assert len(df) == 1
        assert pd.isna(df["beta"].iloc[0])

    def test_error_ticker_skipped(self):
        from factor_engine import compute_metrics
        raw = [{"Ticker": "BAD", "_error": "timeout"}]
        df = compute_metrics(raw, pd.Series(dtype=float))
        assert len(df) == 1
        assert df.get("_skipped", pd.Series([False])).iloc[0] == True


class TestCacheRoundTrip:
    def test_parquet_roundtrip(self, scored_df):
        from factor_engine import write_scores_parquet, _find_latest_cache
        write_scores_parquet(scored_df)
        path, _ = _find_latest_cache("factor_scores")
        assert path is not None
        loaded = pd.read_parquet(str(path))
        assert loaded.shape[0] == scored_df.shape[0]
        assert abs(loaded["Composite"].sum() - scored_df["Composite"].sum()) < 1.0


# ===========================================================================
# UNIT TESTS: portfolio_constructor.py
# ===========================================================================

class TestConstructPortfolio:
    def test_correct_count(self, scored_df, cfg):
        from portfolio_constructor import construct_portfolio
        port = construct_portfolio(scored_df, cfg)
        assert len(port) == cfg["portfolio"]["num_stocks"]

    def test_sector_cap_respected(self, scored_df, cfg):
        from portfolio_constructor import construct_portfolio
        port = construct_portfolio(scored_df, cfg)
        max_sec = cfg["portfolio"]["max_sector_concentration"]
        for sec, cnt in port["Sector"].value_counts().items():
            assert cnt <= max_sec, f"Sector {sec} has {cnt} > {max_sec}"

    def test_weights_sum_100(self, scored_df, cfg):
        from portfolio_constructor import construct_portfolio
        port = construct_portfolio(scored_df, cfg)
        assert abs(port["Equal_Weight_Pct"].sum() - 100) < 0.5
        assert abs(port["RiskParity_Weight_Pct"].sum() - 100) < 0.5

    def test_no_nan_weights(self, scored_df, cfg):
        from portfolio_constructor import construct_portfolio
        port = construct_portfolio(scored_df, cfg)
        assert port["Equal_Weight_Pct"].notna().all()
        assert port["RiskParity_Weight_Pct"].notna().all()

    def test_empty_universe(self, scored_df, cfg):
        from portfolio_constructor import construct_portfolio
        empty = scored_df[scored_df["Composite"] > 999].copy()
        port = construct_portfolio(empty, cfg)
        assert port.empty

    def test_zero_volatility(self, scored_df, cfg):
        from portfolio_constructor import construct_portfolio
        df_zv = scored_df.copy()
        df_zv["volatility"] = 0.0
        port = construct_portfolio(df_zv, cfg)
        assert port["RiskParity_Weight_Pct"].notna().all()
        assert np.isfinite(port["RiskParity_Weight_Pct"]).all()
        assert abs(port["RiskParity_Weight_Pct"].sum() - 100) < 0.5

    def test_nan_volatility(self, scored_df, cfg):
        from portfolio_constructor import construct_portfolio
        df_nv = scored_df.copy()
        df_nv["volatility"] = np.nan
        port = construct_portfolio(df_nv, cfg)
        assert port["RiskParity_Weight_Pct"].notna().all()


class TestPortfolioStats:
    def test_has_required_keys(self, scored_df, cfg):
        from portfolio_constructor import construct_portfolio, compute_portfolio_stats
        port = construct_portfolio(scored_df, cfg)
        stats = compute_portfolio_stats(port, cfg)
        for key in ["n_stocks", "avg_composite", "avg_beta", "est_div_yield",
                     "sector_alloc", "factor_exposure", "date_generated"]:
            assert key in stats

    def test_empty_portfolio_stats(self, cfg):
        from portfolio_constructor import compute_portfolio_stats
        empty = pd.DataFrame(columns=["Ticker", "Sector", "Composite",
                                       "Equal_Weight_Pct", "RiskParity_Weight_Pct"])
        stats = compute_portfolio_stats(empty, cfg)
        assert stats["n_stocks"] == 0


# ===========================================================================
# UNIT TESTS: run_screener.py utilities
# ===========================================================================

class TestDQLog:
    def test_log_appends(self):
        from run_screener import _DQ_LOG_ROWS, dq_log
        _DQ_LOG_ROWS.clear()
        dq_log("TEST", "test_type", "Low", "desc", "action")
        assert len(_DQ_LOG_ROWS) == 1
        assert _DQ_LOG_ROWS[0]["Ticker"] == "TEST"

    def test_flush_writes_csv(self, tmp_path):
        from run_screener import _DQ_LOG_ROWS, dq_log, flush_dq_log, VALIDATION_DIR
        _DQ_LOG_ROWS.clear()
        dq_log("A", "fetch_failure", "High", "test", "excluded")
        dq_log("B", "missing_metric", "Medium", "test", "median")
        path, n = flush_dq_log()
        assert n == 2
        assert Path(path).exists()

    def test_clear_on_new_run(self):
        from run_screener import _DQ_LOG_ROWS, dq_log
        dq_log("STALE", "old", "Low", "leftover", "none")
        assert len(_DQ_LOG_ROWS) > 0
        _DQ_LOG_ROWS.clear()
        assert len(_DQ_LOG_ROWS) == 0


class TestLoadConfigSafe:
    def test_loads_successfully(self):
        from run_screener import load_config_safe
        cfg = load_config_safe()
        assert isinstance(cfg, dict)
        assert "factor_weights" in cfg


# ===========================================================================
# INTEGRATION TEST: full pipeline end-to-end
# ===========================================================================

class TestFullPipeline:
    def test_end_to_end(self, scored_df, cfg):
        """Verify the full pipeline produces valid output."""
        from portfolio_constructor import (
            construct_portfolio, compute_portfolio_stats, write_full_excel
        )
        from factor_engine import write_scores_parquet, ROOT

        # Portfolio
        port = construct_portfolio(scored_df, cfg)
        assert len(port) == 25

        # Stats
        stats = compute_portfolio_stats(port, cfg)
        assert stats["n_stocks"] == 25
        assert 0 < stats["avg_composite"] <= 100
        assert 0 <= stats["avg_beta"] <= 3

        # Excel
        excel_path = write_full_excel(scored_df, port, stats, cfg)
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        assert set(wb.sheetnames) == {"FactorScores", "ScreenerDashboard", "ModelPortfolio"}

        # FactorScores sheet
        ws = wb["FactorScores"]
        assert ws.cell(1, 1).value == "Ticker"
        assert ws.max_row > 100  # 392 data rows + header

        # ScreenerDashboard
        ws2 = wb["ScreenerDashboard"]
        assert "SCREENER DASHBOARD" in str(ws2.cell(1, 1).value)

        # ModelPortfolio
        ws3 = wb["ModelPortfolio"]
        assert "MODEL PORTFOLIO" in str(ws3.cell(1, 1).value)

        # Parquet
        pq_path = write_scores_parquet(scored_df)
        loaded = pd.read_parquet(pq_path)
        assert len(loaded) == len(scored_df)

    def test_tiny_pipeline(self, cfg):
        """3-ticker pipeline runs without crash."""
        from factor_engine import (
            get_sp500_tickers, _generate_sample_data,
            winsorize_metrics, compute_sector_percentiles,
            compute_category_scores, compute_composite,
            apply_value_trap_flags, rank_stocks,
        )
        from portfolio_constructor import construct_portfolio, compute_portfolio_stats

        univ = get_sp500_tickers(cfg).head(3)
        df = _generate_sample_data(univ, seed=99)
        df = winsorize_metrics(df)
        df = compute_sector_percentiles(df)

        import copy
        _cfg = copy.deepcopy(cfg)
        _cfg["factor_weights"]["revisions"] = 0
        others = [k for k in _cfg["factor_weights"] if k != "revisions"]
        s = sum(_cfg["factor_weights"][k] for k in others)
        for k in others:
            _cfg["factor_weights"][k] *= 100 / s

        df = compute_category_scores(df, _cfg)
        df = compute_composite(df, _cfg)
        df = apply_value_trap_flags(df, _cfg)
        df = rank_stocks(df)
        assert len(df) == 3

        port = construct_portfolio(df, _cfg)
        assert len(port) > 0
        stats = compute_portfolio_stats(port, _cfg)
        assert stats["n_stocks"] > 0


# ===========================================================================
# Tests for Fix #1: _stmt_val exact matching
# ===========================================================================
class TestStmtVal:
    def test_exact_match_preferred(self):
        from factor_engine import _stmt_val
        data = {"2024": [100.0, 200.0], "2023": [90.0, 180.0]}
        stmt = pd.DataFrame(data, index=[
            "Net Income From Continuing Operations",
            "Net Income",
        ])
        result = _stmt_val(stmt, "Net Income")
        assert result == 200.0

    def test_substring_fallback(self):
        from factor_engine import _stmt_val
        data = {"2024": [50.0]}
        stmt = pd.DataFrame(data, index=["Total Revenue Adjusted"])
        result = _stmt_val(stmt, "Total Revenue")
        assert result == 50.0

    def test_empty_stmt(self):
        from factor_engine import _stmt_val
        result = _stmt_val(None, "Net Income")
        assert np.isnan(result)


# ===========================================================================
# Tests for Fix #2: Beta date alignment
# ===========================================================================
class TestBetaDateAlignment:
    def test_dict_format_beta(self):
        from factor_engine import compute_metrics
        dates = pd.date_range("2024-01-01", periods=252, freq="B")
        rng = np.random.default_rng(42)
        market_returns = pd.Series(rng.normal(0, 0.01, 252), index=dates)
        dr_dict = {d.strftime("%Y-%m-%d"): rng.normal(0, 0.012) for d in dates}
        raw = [{"Ticker": "TEST", "marketCap": 1e12, "enterpriseValue": 1e12,
                "sector": "Tech", "shortName": "Test Co",
                "_daily_returns": dr_dict, "volatility_1y": 0.20}]
        df = compute_metrics(raw, market_returns)
        assert pd.notna(df["beta"].iloc[0])

    def test_list_format_backward_compat(self):
        from factor_engine import compute_metrics
        dates = pd.date_range("2024-01-01", periods=252, freq="B")
        rng = np.random.default_rng(42)
        market_returns = pd.Series(rng.normal(0, 0.01, 252), index=dates)
        raw = [{"Ticker": "TEST", "marketCap": 1e12, "enterpriseValue": 1e12,
                "sector": "Tech", "shortName": "Test Co",
                "_daily_returns": list(rng.normal(0, 0.012, 252)),
                "volatility_1y": 0.20}]
        df = compute_metrics(raw, market_returns)
        assert pd.notna(df["beta"].iloc[0])


# ===========================================================================
# Tests for Fix #3: Piotroski normalization
# ===========================================================================
class TestPiotroskiNormalization:
    def test_full_data(self):
        from factor_engine import compute_metrics
        raw = [{"Ticker": "PIO", "marketCap": 1e12, "enterpriseValue": 1e12,
                "sector": "Industrials", "shortName": "Pio Inc",
                "netIncome": 100, "netIncome_prior": 80,
                "operatingCashFlow": 120,
                "totalAssets": 1000, "totalAssets_prior": 950,
                "longTermDebt": 200, "longTermDebt_prior": 250,
                "currentAssets": 500, "currentAssets_prior": 480,
                "currentLiabilities": 300, "currentLiabilities_prior": 310,
                "sharesBS": 100, "sharesBS_prior": 100,
                "grossProfit": 400, "grossProfit_prior": 370,
                "totalRevenue": 800, "totalRevenue_prior": 750}]
        df = compute_metrics(raw, pd.Series(dtype=float))
        f = df["piotroski_f_score"].iloc[0]
        assert pd.notna(f) and 0 <= f <= 9

    def test_partial_data_normalized(self):
        from factor_engine import compute_metrics
        raw = [{"Ticker": "PAR", "marketCap": 1e12, "enterpriseValue": 1e12,
                "sector": "Tech", "shortName": "Partial",
                "netIncome": 100, "operatingCashFlow": 120,
                "totalAssets": 1000,
                "sharesBS": 100, "sharesBS_prior": 100}]
        df = compute_metrics(raw, pd.Series(dtype=float))
        f = df["piotroski_f_score"].iloc[0]
        assert pd.notna(f) and 0 <= f <= 9

    def test_too_few_signals_is_nan(self):
        from factor_engine import compute_metrics
        raw = [{"Ticker": "FEW", "marketCap": 1e12, "enterpriseValue": 1e12,
                "sector": "Tech", "shortName": "Few Data",
                "netIncome": 100}]
        df = compute_metrics(raw, pd.Series(dtype=float))
        assert pd.isna(df["piotroski_f_score"].iloc[0])


# ===========================================================================
# Tests for Fix #4: ROIC formula
# ===========================================================================
class TestROICFormula:
    def test_standard_ic_formula(self):
        from factor_engine import compute_metrics
        raw = [{"Ticker": "IC", "marketCap": 1e12, "enterpriseValue": 1e12,
                "sector": "Tech", "shortName": "IC Test",
                "ebit": 100, "incomeTaxExpense": 21, "pretaxIncome": 100,
                "totalEquity": 500, "totalDebt": 200, "totalCash": 50,
                "totalAssets": 1000}]
        df = compute_metrics(raw, pd.Series(dtype=float))
        expected = 79 / 650  # NOPAT=79, IC=500+200-50=650
        assert abs(df["roic"].iloc[0] - expected) < 0.001


# ===========================================================================
# Tests for Fix #5: Drift alerts
# ===========================================================================
class TestDriftAlerts:
    def test_drift_alert_threshold(self):
        from run_screener import _DQ_LOG_ROWS, dq_log
        _DQ_LOG_ROWS.clear()
        stats_missing = {"ROIC": 60.0, "Beta": 30.0, "Analyst Surprise": 80.0}
        threshold = 50
        drift_alerts = 0
        for lbl, pct in stats_missing.items():
            if pct > threshold:
                dq_log("UNIVERSE", "metric_drift", "High",
                       f"{lbl} missing {pct:.1f}%", "Flagged")
                drift_alerts += 1
        assert drift_alerts == 2
        assert len(_DQ_LOG_ROWS) == 2
        _DQ_LOG_ROWS.clear()


# ===========================================================================
# Tests for Fix #6: Backtest helpers
# ===========================================================================
class TestPerfMetrics:
    def test_empty_returns(self):
        from backtest import _perf_metrics
        result = _perf_metrics([])
        assert result["ann_return"] == 0
        assert result["sharpe"] == 0

    def test_positive_returns(self):
        from backtest import _perf_metrics
        result = _perf_metrics([0.01] * 12)
        assert result["ann_return"] > 0.10

    def test_negative_returns(self):
        from backtest import _perf_metrics
        result = _perf_metrics([-0.05] * 12)
        assert result["ann_return"] < 0
        assert result["max_dd"] < 0


class TestAssignDeciles:
    def test_correct_range(self):
        from backtest import _assign_deciles
        deciles = _assign_deciles(pd.Series(range(100)))
        assert deciles.min() >= 1 and deciles.max() <= 10

    def test_small_input(self):
        from backtest import _assign_deciles
        deciles = _assign_deciles(pd.Series(range(10)))
        assert len(deciles) == 10


class TestMonthlyForwardReturns:
    def test_shape_and_last_row_nan(self):
        from backtest import _monthly_forward_returns
        dates = pd.date_range("2024-01-01", periods=12, freq="ME")
        prices = pd.DataFrame({"A": range(100, 112), "B": range(50, 62)}, index=dates)
        fwd = _monthly_forward_returns(prices)
        assert fwd.shape == prices.shape
        assert fwd.iloc[-1].isna().all()


# ===========================================================================
# Tests for Fix #7: Sector-relative composite
# ===========================================================================
class TestSectorRelativeComposite:
    def test_sector_relative(self, sample_df, cfg):
        import copy
        from factor_engine import (
            winsorize_metrics, compute_sector_percentiles,
            compute_category_scores, compute_composite,
        )
        _cfg = copy.deepcopy(cfg)
        _cfg["factor_weights"]["revisions"] = 0
        others = [k for k in _cfg["factor_weights"] if k != "revisions"]
        s = sum(_cfg["factor_weights"][k] for k in others)
        for k in others:
            _cfg["factor_weights"][k] *= 100 / s
        _cfg.setdefault("sector_neutral", {})["sector_relative_composite"] = True

        df = winsorize_metrics(sample_df.copy())
        df = compute_sector_percentiles(df)
        df = compute_category_scores(df, _cfg)
        df = compute_composite(df, _cfg)

        for sec, grp in df.groupby("Sector"):
            if len(grp) >= 3:
                assert grp["Composite"].min() < 5, f"Sector {sec} min too high"
                assert grp["Composite"].max() > 95, f"Sector {sec} max too low"

    def test_default_is_global(self, sample_df, cfg):
        import copy
        from factor_engine import (
            winsorize_metrics, compute_sector_percentiles,
            compute_category_scores, compute_composite,
        )
        _cfg = copy.deepcopy(cfg)
        _cfg["factor_weights"]["revisions"] = 0
        others = [k for k in _cfg["factor_weights"] if k != "revisions"]
        s = sum(_cfg["factor_weights"][k] for k in others)
        for k in others:
            _cfg["factor_weights"][k] *= 100 / s

        df = winsorize_metrics(sample_df.copy())
        df = compute_sector_percentiles(df)
        df = compute_category_scores(df, _cfg)
        df = compute_composite(df, _cfg)

        assert df["Composite"].min() < 2
        assert df["Composite"].max() > 98


# ===========================================================================
# Tests for Fix #8: Smart retry
# ===========================================================================
class TestSmartRetry:
    def test_non_retryable_pattern_detected(self):
        from factor_engine import _NON_RETRYABLE_PATTERNS
        err = "404 Client Error: Not Found for url"
        assert any(p in err.lower() for p in _NON_RETRYABLE_PATTERNS)

    def test_retryable_error_not_flagged(self):
        from factor_engine import _NON_RETRYABLE_PATTERNS
        err = "Connection timed out after 30s"
        assert not any(p in err.lower() for p in _NON_RETRYABLE_PATTERNS)


# ===========================================================================
# Tests for PEG Ratio
# ===========================================================================
class TestPEGRatio:
    def test_peg_computed(self):
        from factor_engine import compute_metrics
        # P/E = 100/5 = 20, earningsGrowth = 0.25 (25%), PEG = 20 / 25 = 0.8
        raw = [{"Ticker": "PEG", "marketCap": 1e12, "enterpriseValue": 1e12,
                "sector": "Tech", "shortName": "PEG Test",
                "currentPrice": 100, "trailingEps": 5.0,
                "forwardEps": 6.0, "earningsGrowth": 0.25}]
        df = compute_metrics(raw, pd.Series(dtype=float))
        peg = df["peg_ratio"].iloc[0]
        assert pd.notna(peg)
        expected = 20.0 / (0.25 * 100)  # 20 / 25 = 0.8
        assert abs(peg - expected) < 0.01

    def test_peg_nan_for_negative_growth(self):
        from factor_engine import compute_metrics
        # earningsGrowth negative => PEG should be NaN
        raw = [{"Ticker": "NEG", "marketCap": 1e12, "enterpriseValue": 1e12,
                "sector": "Tech", "shortName": "Neg Growth",
                "currentPrice": 100, "trailingEps": 5.0,
                "forwardEps": 4.0, "earningsGrowth": -0.10}]
        df = compute_metrics(raw, pd.Series(dtype=float))
        assert pd.isna(df["peg_ratio"].iloc[0])


# ===========================================================================
# Standalone runner
# ===========================================================================
if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v", "--tb=short"]))
